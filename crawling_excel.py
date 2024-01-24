# %% 크롤링해서 엑셀 파일로 만들기
import openpyxl

excel_file = openpyxl.Workbook()

excel_sheet = excel_file.active
# excel_sheet.title = '리포트'

excel_sheet.append(['data1', 'data2', 'data3'])

excel_file.save('tmp.xlsx')

excel_file.close()

# %% 엑셀 파일 만들기
import openpyxl

def write_excel_template(filename, sheetname, listdata):
    excel_file = openpyxl.Workbook()
    excel_sheet = excel_file.active
    excel_sheet.column_dimensions['A'].width = 100
    excel_sheet.column_dimensions['B'].width = 20

    if sheetname != '':
        excel_sheet.title = sheetname

    for item in listdata:
        excel_sheet.append(item)
    excel_file.save(filename)
    excel_file.close()

# %%
import requests
from bs4 import BeautifulSoup

product_lists = list()

for page_num in range(10):
    if page_num == 0:
        res = requests.get('https://davelee-fun.github.io/')
    else:
        res = requests.get('https://davelee-fun.github.io/page' + str(page_num + 1))
    soup = BeautifulSoup(res.content, 'html.parser')

    data = soup.select('div.card') 

    for item in data:
        product_name = item.select_one('div.card-body h4.card-text')
        product_date = item.select_one('div.wrapfooter span.post-date')
        product_info = [product_name.get_text().strip(), product_date.get_text()]
        product_lists.append(product_info)

# %%
write_excel_template('product.xlsx', '상품정보', product_lists)
    
# %% 엑셀 파일 읽기
import openpyxl

excel_file = openpyxl.load_workbook('product.xlsx') # 엑셀 파일 오픈

excel_file.sheetnames # 해당 엑셀 파일 안에 있는 쉬트 이름 확인

excel_sheet = excel_file["상품정보"]

for item in excel_sheet.rows:
    print(item[0].value, item[1].value)

excel_file.close()
# %%
